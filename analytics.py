import os
import pandas as pd
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import plotly.express as px
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

# ---------------- CONFIG ----------------
from config import DB_URI  # assume config.py stores DB connection string

engine = create_engine(DB_URI)

os.makedirs("charts", exist_ok=True)
os.makedirs("exports", exist_ok=True)

# ---------------- HELPER FUNCTIONS ----------------
def run_query(query: str):
    with engine.connect() as conn:
        return pd.read_sql(text(query), conn)

def save_chart(df, chart_type, filename, title, xlabel=None, ylabel=None):
    plt.figure(figsize=(10, 7))
    
    if chart_type == "pie":
        df.plot.pie(y=df.columns[1], labels=df[df.columns[0]], autopct='%1.1f%%', legend=False)
        plt.ylabel("")
        plt.title(title, fontsize=14, weight="bold")

    elif chart_type == "bar":
        df.plot.bar(x=df.columns[0], y=df.columns[1])
        plt.title(title, fontsize=14, weight="bold")

    elif chart_type == "barh":
        df.plot.barh(x=df.columns[0], y=df.columns[1])
        plt.title(title, fontsize=14, weight="bold")

    elif chart_type == "line":
        ax = df.plot.line(x=df.columns[0], y=df.columns[1], marker="o")
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b-%d"))
        plt.xticks(rotation=45, ha="right")
        years = pd.to_datetime(df[df.columns[0]], errors="coerce").dt.year.dropna().unique()
        year_str = ", ".join(str(y) for y in years)
        plt.title(title or f"Flights per Day (Year {year_str})", fontsize=14, weight="bold")

    elif chart_type == "hist":
        # Changed to bar instead of hist, since it's categorical data (city vs counts)
        try:
            df[df.columns[1]] = pd.to_numeric(df[df.columns[1]], errors='coerce')
            df = df.dropna(subset=[df.columns[1]])
            df.plot.bar(x=df.columns[0], y=df.columns[1])
        except Exception as e:
            print(f"Error generating histogram: {e}")
            return
        plt.title(title, fontsize=14, weight="bold")

    elif chart_type == "scatter":
        short_labels = {name: f"M{i+1}" for i, name in enumerate(df[df.columns[0]].unique())}
        df["short_model"] = df[df.columns[0]].map(short_labels)
        df.plot.scatter(x="short_model", y=df.columns[1])
        plt.xticks(rotation=45, ha="right")
        plt.xlabel(xlabel or "Aircraft Model (M1, M2, …)")
        plt.ylabel(ylabel or "Number of Passengers")
        legend_text = "Model Codes:\n" + "\n".join([f"{v} = {k}" for k, v in short_labels.items()])
        plt.gcf().text(1.02, 0.5, legend_text, fontsize=9, va="center", 
                      bbox=dict(facecolor="white", alpha=0.7))
        plt.title(title, fontsize=14, weight="bold")
    
    if xlabel:
        plt.xlabel(xlabel)
    if ylabel:
        plt.ylabel(ylabel)

    filepath = f"charts/{filename}.png"
    plt.savefig(filepath, bbox_inches="tight")
    plt.close()
    print(f"[Chart Saved] {filename}.png — {len(df)} rows — {chart_type} — {title}")

# ---------------- CUSTOM QUERIES ----------------
queries = {
    "pie": """ 
        SELECT a.city AS destination_city, COUNT(*) AS total_passengers 
        FROM ticket_flights tf 
        JOIN flights f ON tf.flight_id = f.flight_id 
        JOIN airports_data a ON f.arrival_airport = a.airport_code 
        GROUP BY a.city 
        ORDER BY total_passengers DESC LIMIT 6;
    """,
    "bar": """ 
        SELECT a.model AS aircraft_model, COUNT(*) AS usage_count 
        FROM flights f 
        JOIN aircrafts_data a ON f.aircraft_code = a.aircraft_code 
        GROUP BY a.model 
        ORDER BY usage_count DESC LIMIT 6;
    """,
    "barh": """ 
        SELECT a.model AS aircraft_model, COUNT(s.seat_no) AS total_seats 
        FROM seats s 
        JOIN aircrafts_data a ON s.aircraft_code = a.aircraft_code 
        GROUP BY a.model 
        ORDER BY total_seats DESC LIMIT 6;
    """,
    "line": """ 
        SELECT DATE(scheduled_departure) AS dep_date, COUNT(*) AS total_flights 
        FROM flights 
        WHERE scheduled_departure IS NOT NULL 
        GROUP BY dep_date 
        ORDER BY dep_date 
        LIMIT 30;
    """,
    "hist": """ 
        SELECT 
            a.city AS destination_city,
            COUNT(*) AS business_class_passenger_count
        FROM ticket_flights tf
        JOIN flights f ON tf.flight_id = f.flight_id
        JOIN airports_data a ON f.arrival_airport = a.airport_code
        WHERE tf.fare_conditions = 'Business'
        GROUP BY a.city
        ORDER BY business_class_passenger_count DESC
        LIMIT 30;
    """,
    "scatter": """ 
        SELECT ad.model AS aircraft_model, COUNT(tf.ticket_no) AS passengers 
        FROM ticket_flights tf 
        JOIN flights f ON tf.flight_id = f.flight_id 
        JOIN aircrafts_data ad ON f.aircraft_code = ad.aircraft_code 
        GROUP BY ad.model 
        ORDER BY passengers DESC LIMIT 15;
    """
}

# ---------------- TITLES ----------------
titles = {
    "pie": "Hottest Destinations by Number of Passengers",
    "bar": "Top 5 Most Used Aircraft Models",
    "barh": "Top 5 Aircraft Models with the Highest Number of Seats",
    "line": "Flights per Day (Time Series)",
    "hist": "Top Destinations with the Highest Number of Business Class Passengers",
    "scatter": "Aircraft Model vs Total Passengers"
}

labels = {
    "pie": (None, None),
    "bar": ("Aircraft Model", "Usage Count"),
    "barh": ("Total Seats", "Aircraft Model"),
    "line": ("Date", "Total Flights"),
    "hist": ("Destination City", "Business Class Passenger Count"),
    "scatter": ("Aircraft Model", "Passengers")
}

# ---------------- CHART GENERATION ----------------
for chart_type, query in queries.items():
    df = run_query(query)
    save_chart(
        df, 
        chart_type, 
        chart_type, 
        titles.get(chart_type, "Auto Chart"),
        labels.get(chart_type, (None, None))[0],
        labels.get(chart_type, (None, None))[1]
    )

# ---------------- INTERACTIVE PLOTLY (Weekly Slider with Lines) ----------------
df_time = run_query(""" 
    SELECT DATE(scheduled_departure) AS dep_date, COUNT(*) AS total_flights 
    FROM flights 
    WHERE scheduled_departure IS NOT NULL 
    GROUP BY dep_date 
    ORDER BY dep_date 
    LIMIT 100;
""")

import plotly.graph_objects as go
import pandas as pd

# Ensure dates are datetime
df_time["dep_date"] = pd.to_datetime(df_time["dep_date"])

# Group into weeks (7-day intervals starting Monday)
df_time["week_start"] = df_time["dep_date"] - pd.to_timedelta(df_time["dep_date"].dt.weekday, unit="d")
weeks = df_time["week_start"].unique()

# Create base figure
fig = go.Figure()

# First week's data
first_week = weeks[0]
week_df = df_time[df_time["week_start"] == first_week]

fig.add_trace(go.Scatter(
    x=week_df["dep_date"],
    y=week_df["total_flights"],
    mode="lines+markers",   # <--- dots + connecting lines
    line=dict(color="blue", width=2),
    marker=dict(size=8),
    name="Flights"
))

# Frames for each week
frames = []
for wk in weeks:
    week_df = df_time[df_time["week_start"] == wk]
    frames.append(go.Frame(
        data=[go.Scatter(
            x=week_df["dep_date"],
            y=week_df["total_flights"],
            mode="lines+markers",   # <--- dots + connecting lines
            line=dict(color="blue", width=2),
            marker=dict(size=8)
        )],
        name=str(wk.date())
    ))

# Slider steps
steps = []
for wk in weeks:
    step = dict(
        method="animate",
        args=[[str(wk.date())],
              {"mode": "immediate", "frame": {"duration": 500, "redraw": True},
               "transition": {"duration": 0}}],
        label=str(wk.date())
    )
    steps.append(step)

sliders = [dict(
    active=0,
    currentvalue={"prefix": "Week starting: "},
    pad={"t": 50},
    steps=steps
)]

# Layout
fig.update_layout(
    title="Flights Over Time (Weekly View with Lines)",
    sliders=sliders,
    xaxis_title="Date",
    yaxis_title="Total Flights",
    yaxis=dict(dtick=50),  # keep tick spacing larger
    updatemenus=[dict(type="buttons", showactive=False,
                      buttons=[dict(label="Play",
                                    method="animate",
                                    args=[None, {"frame": {"duration": 800, "redraw": True},
                                                 "fromcurrent": True}]),
                               dict(label="Pause",
                                    method="animate",
                                    args=[[None], {"frame": {"duration": 0, "redraw": False},
                                                   "mode": "immediate"}])])]
)

fig.frames = frames

fig.show()


# ---------------- EXPORT TO EXCEL ----------------
def export_to_excel(dataframes_dict, filename):
    filepath = f"exports/{filename}"
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            # use descriptive titles as sheet names
            df.to_excel(writer, sheet_name=titles.get(sheet_name, sheet_name), index=False)
    
    wb = load_workbook(filepath)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.iter_cols(min_row=2, min_col=2):
            col_letter = col[0].column_letter
            max_row = ws.max_row
            rule = ColorScaleRule(
                start_type="min", start_color="FFAA0000", 
                mid_type="percentile", mid_value=50, mid_color="FFFFFF00", 
                end_type="max", end_color="FF00AA00"
            )
            ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{max_row}", rule)
    
    wb.save(filepath)
    print(f"[Excel Export] Created file {filename}, {len(dataframes_dict)} sheets")

# Example export
export_to_excel({name: run_query(q) for name, q in queries.items()}, "report.xlsx")
